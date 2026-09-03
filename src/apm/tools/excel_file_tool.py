"""Excel connector for workbooks that are actual .xlsx files sitting
locally on disk or stored as files in Google Drive — distinct from
`excel_tool.py`, which only operates on workbooks already living on
OneDrive/SharePoint via Microsoft Graph and has no concept of a local
file or a plain Drive-hosted one.

Both a local file and a Google Drive file are just a source of .xlsx
bytes to this connector: `WorkbookSource` is the abstraction (parallel to
excel_tool.ExcelClient), with `LocalWorkbookSource` and
`GoogleDriveWorkbookSource` as the two implementations. The bytes are
parsed/edited with openpyxl — there's no cell-range API to call for
either source, unlike Graph's Excel API.

Read (list worksheets, read a range) and write (overwrite a range) are
both implemented here, following the same dry_run-gated pattern as
excel_tool.ExcelTool.write_range: write_range defaults to dry_run=True
and must only ever be called with dry_run=False from inside the agent
graph's execute_node, after an approved human interrupt (see
.claude/skills/tool-integration/SKILL.md). It is not yet wired into
apm.agent.graph — same status as the MS Excel connector's write_range.

Known gap: read_range loads with data_only=True to return each formula
cell's last-calculated value rather than the formula text, but openpyxl
never calculates formulas itself — a workbook that has never been opened
in real Excel/Sheets (e.g. one created purely by openpyxl) has no cached
value and such cells read back as None.

Google Drive scope note (found by live testing, not just reasoned
about): this originally requested the narrow `drive.file` scope for
read/write, on the least-privilege theory that it's the smallest scope
covering both. In practice `drive.file` only grants access to files the
app itself created or that the user explicitly opened via a Google
Picker with this app — neither is true for an existing file a user just
points the connector at by id/name. Reads against such a file happened
to succeed anyway (unclear why — possibly an inconsistency in how Drive
enforces the scope for `alt=media` downloads specifically), but writes
failed outright with a 403 `insufficientPermissions`
("Request had insufficient authentication scopes"). Since there's no
Picker flow here, GOOGLE_DRIVE_SCOPE below is the full `drive` scope
instead — broader than ideal, but `drive.file` is not usable for this
connector's actual use case (an existing file, not one the app created).
"""

from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Any, Protocol

from apm.config import Settings
from apm.state.store import StateStore
from apm.tools._retry import with_retry
from apm.tools.base import ActionResult, BaseTool, Capability

# Full Drive read/write access -- see the module docstring's "Google
# Drive scope note" for why the narrower drive.file scope doesn't work
# here (confirmed by a live 403 on write, not just reasoned about).
GOOGLE_DRIVE_SCOPE = "https://www.googleapis.com/auth/drive"

EXCEL_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class WorkbookSource(Protocol):
    """Where the .xlsx bytes live. ExcelFileTool only ever reads/writes
    whole-workbook bytes through this — it doesn't know or care whether
    that's a local path or a Drive file. Both the real sources below and
    test fakes implement just this.
    """

    def describe(self) -> str: ...

    def read_bytes(self) -> bytes: ...

    def write_bytes(self, data: bytes) -> None: ...


@dataclass(frozen=True)
class RangeData:
    sheet_name: str
    address: str
    values: list[list[Any]] = field(default_factory=list)


class ExcelFileTool(BaseTool):
    """Excel connector bound to one workbook file (local or Google
    Drive) at construction time: reads worksheets/ranges, and writes a
    range — the latter only ever reachable through the agent's approval
    interrupt (see apm.agent.graph), same as excel_tool.ExcelTool.
    """

    name = "excel_file"
    capabilities = frozenset({Capability.READ, Capability.WRITE})

    def __init__(self, state: StateStore, source: WorkbookSource) -> None:
        super().__init__(state)
        self._source = source

    def health_check(self) -> bool:
        try:
            self._load_workbook(data_only=True)
            return True
        except Exception:
            return False

    def list_worksheets(self, process_id: str) -> list[str]:
        workbook = self._load_workbook(data_only=True)
        names = list(workbook.sheetnames)
        self._log(
            process_id,
            "read",
            f"Listed {len(names)} worksheet(s) from {self._source.describe()}",
            {"worksheets": names, "source": self._source.describe()},
        )
        return names

    def read_range(
        self, process_id: str, sheet_name: str | None = None, address: str | None = None
    ) -> RangeData:
        """Read a cell range and return its values as a list of rows.

        Only the tool's workbook (set at construction time, e.g. a file
        name via build_local_excel_tool) is required to read something
        useful: `sheet_name` defaults to the workbook's first worksheet,
        and `address` defaults to that worksheet's whole used range
        (openpyxl's `dimensions`) — pass them explicitly (e.g.
        sheet_name="Renewals", address="A1:D20") to read something more
        specific.
        """
        workbook = self._load_workbook(data_only=True)
        resolved_sheet = sheet_name if sheet_name is not None else workbook.sheetnames[0]
        worksheet = workbook[resolved_sheet]
        resolved_address = address if address is not None else worksheet.dimensions
        values = self._range_values(worksheet, resolved_address)
        data = RangeData(sheet_name=resolved_sheet, address=resolved_address, values=values)
        self._log(
            process_id,
            "read",
            f"Read range {resolved_sheet}!{resolved_address} ({len(values)} row(s)) from {self._source.describe()}",
            {
                "sheet_name": resolved_sheet,
                "address": resolved_address,
                "row_count": len(values),
                "source": self._source.describe(),
            },
        )
        return data

    def write_range(
        self, process_id: str, sheet_name: str, address: str, values: list[list[Any]], dry_run: bool = True
    ) -> ActionResult:
        """Overwrite a cell range with new values (a list of rows). Only
        ever call this with dry_run=False from inside the agent graph,
        after the human-approval interrupt has returned an approval.
        """
        summary = f"Write {len(values)} row(s) to {sheet_name}!{address} in {self._source.describe()}"
        self.require_dry_run_guard(dry_run, process_id, summary)

        if dry_run:
            return ActionResult(
                executed=False,
                description=summary,
                details={
                    "sheet_name": sheet_name,
                    "address": address,
                    "values": values,
                    "source": self._source.describe(),
                },
            )

        # data_only=False here (the default): saving back a workbook
        # loaded with data_only=True would permanently replace every
        # other formula cell with its last-cached value.
        workbook = self._load_workbook(data_only=False)
        worksheet = workbook[sheet_name]
        for row_cells, row_values in zip(self._cell_rows(worksheet, address), values):
            for cell, value in zip(row_cells, row_values):
                cell.value = value

        buffer = BytesIO()
        workbook.save(buffer)
        self._source.write_bytes(buffer.getvalue())

        return ActionResult(
            executed=True,
            description=summary,
            details={
                "sheet_name": sheet_name,
                "address": address,
                "row_count": len(values),
                "source": self._source.describe(),
            },
        )

    def _load_workbook(self, data_only: bool):
        from openpyxl import load_workbook

        return load_workbook(BytesIO(self._source.read_bytes()), data_only=data_only)

    @staticmethod
    def _cell_rows(worksheet: Any, address: str) -> list[tuple[Any, ...]]:
        """openpyxl returns a bare Cell for a single-cell address like
        "A1", but a tuple of row-tuples of Cells for a range like
        "A1:B2" — normalize to always be rows of cells.
        """
        cell_range = worksheet[address]
        if hasattr(cell_range, "value"):
            return [(cell_range,)]
        return list(cell_range)

    @classmethod
    def _range_values(cls, worksheet: Any, address: str) -> list[list[Any]]:
        return [[cell.value for cell in row] for row in cls._cell_rows(worksheet, address)]


class LocalWorkbookSource:
    """A workbook stored as a local .xlsx file on disk."""

    def __init__(self, path: Path | str) -> None:
        self._path = Path(path)

    def describe(self) -> str:
        return f"local:{self._path}"

    def read_bytes(self) -> bytes:
        return self._path.read_bytes()

    def write_bytes(self, data: bytes) -> None:
        self._path.write_bytes(data)


class GoogleDriveWorkbookSource:
    """A workbook stored as an .xlsx file in Google Drive, addressed by
    its Drive file id (find it with `scripts/excel_file_demo.py --list`
    or from the file's Drive URL). Imports googleapiclient lazily so
    this module — and ExcelFileTool's unit tests, which use a fake
    source — don't require that dependency at import time.

    Note: this is for an actual .xlsx uploaded to Drive, not a native
    Google Sheet — Sheets is a different file type/API (Sheets v4) with
    no .xlsx bytes to download; out of scope for this connector.
    """

    def __init__(self, credentials: Any, file_id: str) -> None:
        from googleapiclient.discovery import build

        self._service = build("drive", "v3", credentials=credentials)
        self._file_id = file_id

    def describe(self) -> str:
        return f"gdrive:{self._file_id}"

    @with_retry()
    def read_bytes(self) -> bytes:
        from googleapiclient.http import MediaIoBaseDownload

        request = self._service.files().get_media(fileId=self._file_id)
        buffer = BytesIO()
        downloader = MediaIoBaseDownload(buffer, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        return buffer.getvalue()

    # Deliberately NOT retried -- see apm.tools._retry's module docstring:
    # a dropped connection after Drive already wrote the file must not
    # turn into an automatic duplicate write.
    def write_bytes(self, data: bytes) -> None:
        from googleapiclient.http import MediaIoBaseUpload

        media = MediaIoBaseUpload(BytesIO(data), mimetype=EXCEL_MIME_TYPE, resumable=False)
        self._service.files().update(fileId=self._file_id, media_body=media).execute()


def build_local_excel_tool(state: StateStore, path: Path | str) -> ExcelFileTool:
    """Convenience factory for a workbook that's just a file on disk —
    no credentials or network involved at all.
    """
    return ExcelFileTool(state, LocalWorkbookSource(path))


DEFAULT_DRIVE_TOKEN_PATH = Path(".google_drive_token.json")


def build_gdrive_excel_tool(
    state: StateStore, settings: Settings, file_id: str, token_path: Path = DEFAULT_DRIVE_TOKEN_PATH
) -> ExcelFileTool:
    """Convenience factory: runs the Google OAuth flow (if needed,
    reusing the same installed-app flow helper as Gmail/Calendar) and
    returns an ExcelFileTool bound to the given Drive file id.

    Uses its own token cache file (not google_auth.DEFAULT_TOKEN_PATH),
    deliberately — see load_credentials' docstring: a cached token is
    only valid for the scopes it was consented to, so requesting the
    Drive scope alone against the Gmail/Calendar token file would
    overwrite their cached token and break them at request time. Also
    used as-is (same scope, same token file) for
    scripts/excel_file_demo.py's --list/name-lookup helpers, which need
    no narrower a scope than this tool already requests.
    """
    from apm.tools.google_auth import load_credentials

    credentials = load_credentials(settings, scopes=[GOOGLE_DRIVE_SCOPE], token_path=token_path)
    return ExcelFileTool(state, GoogleDriveWorkbookSource(credentials, file_id))
